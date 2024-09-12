import os
import openpyxl
import pandas as pd
import sqlite3
from data_processing import DataProcessing
from processor.folder_manager import FolderManager
from resource_processor.sql_reserved_words import SQLITE_RESERVED_WORDS
from base_class.data_processor_enums import (AvoidUnpackKeysToT2,
                                                    AvoidUnpackKeysToTX,
                                                    DataStructureVar,
                                                    SqlType,
                                                    TxDualIndex,
                                                    TxThirdIndexDimension, 
                                                    UnpackKeysToT2, 
                                                    UnpackKeysToTX)
from resource_processor.utils_decorators import CounterLog
from rich.progress import Progress, TaskID

if __name__ == "__main__":
    from copy import deepcopy
    from resource_processor.hash_file import hash_database

ct_log:CounterLog = CounterLog(print_message=False)

class DuplicatedColumnNames(Exception):
    """Exception raised for duplicated column names."""
    
    def __init__(self, message="", code=100):
        super().__init__(message)
        self.code = code

    def __str__(self):
        # Custom string representation that includes the error message and code
        return f"{self.args[0]} (Error code: {self.code})"
    


class DataStructure(DataProcessing):
    
    def __init__(self, **kwargs) -> None:
    
    # Relationships CASE BD:
    # TX CAN BE MANY TABLES, DYNAMIC CREATED. 
    # T1 (one) --- (many) T2
    # T2 (one) --- (many) TX×
    # SOURCE DATA T1 DICT[STR, ANY]
    # SOURCE DATA T2_TX LIST[DICT[STR, DICT[STR, TX ]]]
    # TX = LIST[DICT]
    # Visualization of relationships:
    # T1
    #  |
    #  +-- T2
    #  |    |
    #  |    +-- TX×
    #  |    |
    #  |    +-- TX×
    #  |
    #  +-- T2
    #       |
    #       +-- TX×
    #       |
    #       +-- TX×

        super().__init__()
        
        
    def init_variables_structure(self,data_structure_var:DataStructureVar):
        # VAR ARGUMENT
        self.DATA_STRUCTURE_VAR:DataStructureVar = data_structure_var
        self.DICT_DATA_EXTRUCTURE_T1:dict[str,any] = data_structure_var.DICT_DATA_EXTRUCTURE_T1
        self.DICT_DATA_EXTRUCTURE_T2_TX:dict[str,any] = data_structure_var.DICT_DATA_EXTRUCTURE_T2_TX
        self.DATA_STRUCTURE_VAR.validate_tx_dual_index()
        # MODELS STRUCTURE
        self.__dict_model_t2:dict[str,any] | None = None
        self.__list_dict_model_tx:list[dict[str,any]] | None = None
        self.__data_frame_model_tx_one_row:pd.DataFrame | None = None
        self.__data_frame_model_tx_empty:pd.DataFrame | None = None
        # MODELS STRUCTURE WITH DATA  
        self.__data_frame_t1:pd.DataFrame | None = None
        self.__data_frame_t2:pd.DataFrame | None = None
        self.__data_frame_tx:pd.DataFrame | None = None
        
    def init_variables_data(self,
                            dict_data_source_t1:dict[str,any],
                            lis_data_source_t2_tx:list[dict[str,any]],):
        #DATA
        self.DICT_DATA_SOURCE_T1:dict[str,any] = dict_data_source_t1
        self.LIST_DATA_SOURCE_T2_TX:list[dict[str,any]]= lis_data_source_t2_tx
        
        
        
    @property
    def dict_model_t2(self) -> dict[str,any]:
        if self.__dict_model_t2 is None:
            self._init_dict_model_t2()
        return self.__dict_model_t2
    
    @property
    def list_dict_model_tx(self) ->list[dict[str,any]]:
        if self.__list_dict_model_tx is None:
            self._init_list_dict_model_tx()
        return self.__list_dict_model_tx
    
    @property
    def data_frame_model_tx_one_row(self) -> pd.DataFrame:
        """
        Use this one to create a database table, 
        the one row is used to define the data type of the columns .
        """
        if self.__data_frame_model_tx_one_row is None:
            self._init_data_frame_tx_one_row()
        return self.__data_frame_model_tx_one_row
    
    @property
    def data_frame_model_tx_empty(self) -> pd.DataFrame:
        if self.__data_frame_model_tx_empty is None:
            self._init_data_frame_tx_empty()
        return self.__data_frame_model_tx_empty
    
    @property
    def data_frame_t1(self) -> pd.DataFrame:
        if self.__data_frame_t1 is None:
            self._init_data_frame_t1()
            #print(self.__data_frame_t1)
        return self.__data_frame_t1
    
    
    @property
    def data_frame_t2(self) -> pd.DataFrame:
        if self.__data_frame_t2 is None:
            self._init_data_frame_t2()
            #print(self.__data_frame_t2)
        return self.__data_frame_t2
    
    
    @property
    def data_frame_tx(self) -> pd.DataFrame:
        if self.__data_frame_tx is None:
            self._init_data_frame_tx()
        return self.__data_frame_tx
    
            
    def _init_dict_model_t2(self)-> None:
        self.__dict_model_t2 = self._build_dictionary(dictionary_data=self.DICT_DATA_EXTRUCTURE_T2_TX,
                                                keys_to_unpack=self.DATA_STRUCTURE_VAR.KEYS_TO_UNPACK_IN_DATA_DICT_T2.KEYS,
                                                keys_to_remove=self.DATA_STRUCTURE_VAR.KEYS_TO_AVOID_UNPACK_DATA_DICT_T2.KEYS)
         
    def _init_list_dict_model_tx(self) -> None:
        
        # LIST DICT USED TO CREATE DYNAMIC TX× . 
        tx_tables_dict = self._filter_and_flattens_dict_from_keywords(input_dict=self.DICT_DATA_EXTRUCTURE_T2_TX,
                                                                            keywords_to_get=[self.DATA_STRUCTURE_VAR.KEY_LIST_TABLE_TX],
                                                                            keywords_to_ignore=self.DATA_STRUCTURE_VAR.KEYS_TO_UNPACK_IN_DATA_DICT_T2.KEYS)
        
        
        dict_to_include_in_txs = self._filter_and_flattens_dict_from_keywords(input_dict=self.DICT_DATA_EXTRUCTURE_T2_TX,
                                                                                    keywords_to_get=self.DATA_STRUCTURE_VAR.KEYS_TO_UNPACK_IN_TX.KEYS,
                                                                                    keywords_to_ignore=self.DATA_STRUCTURE_VAR.KEYS_TO_AVOID_UNPACK_IN_TX.KEYS)
        
        
        
    
        list_model_to_add_tx = self._create_list_to_extend_in_tx(dict_to_include_in_txs=dict_to_include_in_txs, 
                                                                tx_dual_index=self.DATA_STRUCTURE_VAR.TX_DUAL_INDEX,
                                                                tx_third_index_dimension=self.DATA_STRUCTURE_VAR.TX_THIRD_INDEX_DIMENSION)
        
        list_dict_tx_model:list[dict] = tx_tables_dict[self.DATA_STRUCTURE_VAR.KEY_LIST_TABLE_TX]
        # INSERT OPTIONAL COLUMNS IN TX .  
        list_dict_tx_model.extend(list_model_to_add_tx)        
        self.__list_dict_model_tx = list_dict_tx_model
        
    
    def _init_data_frame_tx_one_row(self) -> None: 
        
        self.__data_frame_model_tx_one_row =  self._create_multi_index_data_frame_model(
                                                                     list_dict_model=self.list_dict_model_tx,
                                                                     tx_dual_index=self.DATA_STRUCTURE_VAR.TX_DUAL_INDEX,
                                                                     third_dimension_index_keys=self.DATA_STRUCTURE_VAR.TX_THIRD_INDEX_DIMENSION,
                                                                     shared_foreign_key=self.DATA_STRUCTURE_VAR.SHARED_FOREIGN_KEY_T2_TX,
                                                                     )
    
    def _init_data_frame_tx_empty(self) -> None:
        self.__data_frame_model_tx_empty =  self._create_multi_index_data_frame_model(list_dict_model=self.list_dict_model_tx,
                                                                    tx_dual_index=self.DATA_STRUCTURE_VAR.TX_DUAL_INDEX,
                                                                    third_dimension_index_keys=self.DATA_STRUCTURE_VAR.TX_THIRD_INDEX_DIMENSION,
                                                                    shared_foreign_key=self.DATA_STRUCTURE_VAR.SHARED_FOREIGN_KEY_T2_TX,
                                                                    with_data=False,
                                                                    )
        
    
    def _init_data_frame_t1(self) -> None:
        self.__data_frame_t1 = self._build_data_frame_from_dict(self.DICT_DATA_SOURCE_T1)
    
    
    def _init_data_frame_t2(self) -> None:
        self.__data_frame_t2 = self.create_data_frame_from_model(list_data=self.LIST_DATA_SOURCE_T2_TX,
                                                                keywords_to_get=list(self.dict_model_t2.keys())
                                                                )
    
    def _init_data_frame_tx(self) -> None:
        
        list_shared_keys_ts_tx:list[str] = self.data_frame_t2[self.DATA_STRUCTURE_VAR.PRIMARY_KEY_T2].tolist()
        
        key_list_table_tx:str = self.DATA_STRUCTURE_VAR.KEY_LIST_TABLE_TX
        keywords_to_get:list[str] = self.DATA_STRUCTURE_VAR.KEYS_TO_UNPACK_IN_TX.KEYS
        keywords_to_ignore:list[str] = self.DATA_STRUCTURE_VAR.KEYS_TO_AVOID_UNPACK_IN_TX.KEYS
        tx_dual_index:TxDualIndex = self.DATA_STRUCTURE_VAR.TX_DUAL_INDEX
        tx_third_index_dimension:TxThirdIndexDimension = self.DATA_STRUCTURE_VAR.TX_THIRD_INDEX_DIMENSION
        
        super_list_tx:list[list[dict]] = [self._extend_list_columns_txs(entry_dict=entry,
                                                          key_list_table_tx=key_list_table_tx,
                                                          keywords_to_get=keywords_to_get,
                                                          keywords_to_ignore=keywords_to_ignore,
                                                          tx_dual_index=tx_dual_index,
                                                          tx_third_index_dimension=tx_third_index_dimension,
                                                          ) for entry in self.LIST_DATA_SOURCE_T2_TX]

    

        self.__data_frame_tx = self._populate_data_frame_tx(  super_list_data=super_list_tx,
                                                                    df_model=self.data_frame_model_tx_empty,
                                                                    tx_dual_index=tx_dual_index,
                                                                    third_dimension_index_keys=tx_third_index_dimension,
                                                                    list_keys=list_shared_keys_ts_tx)
    
    
    def export_spreadshet(self, path:str, progress:Progress, task_id:TaskID, has_links:bool = False, adjust_columns:bool= True):

        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            
            self.data_frame_tx.to_excel(writer, sheet_name=self.DATA_STRUCTURE_VAR.TX_DEFAULT_NAME, index=True)
            progress.update(task_id, advance=1)
            
            self.data_frame_t1.to_excel(writer, sheet_name=self.DATA_STRUCTURE_VAR.T1_NAME, index=True)
            progress.update(task_id, advance=1)
            
            self.data_frame_t2.to_excel(writer, sheet_name=self.DATA_STRUCTURE_VAR.T2_NAME, index=True)
            progress.update(task_id, advance=1)
        
        if has_links:
            
            self.export_dataframe_with_links(df=self.data_frame_tx,
                                                path=path,
                                                sheet_name=self.DATA_STRUCTURE_VAR.TX_DEFAULT_NAME)
            progress.update(task_id, advance=1)
        
        if adjust_columns:
            
            self.adjust_column_widths(path)
            progress.update(task_id, advance=1)

    @staticmethod            
    def export_dataframe_with_links(df: pd.DataFrame, path: str, sheet_name:str):

        
        workbook = openpyxl.load_workbook(path)
        sheet = workbook[sheet_name]
        empty_row = 1
        levels_columns = df.columns.nlevels
        levels_columns += empty_row
        
        for row_idx, row in enumerate(df.itertuples(), start=levels_columns +1):
            for col_idx, value in enumerate(row, start=1):  # Iterate through the row values
                
                if isinstance(value, str) and value.startswith('file:///'):
                    link = value
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    cell.value = "link" 
                    cell.hyperlink = link  
                    cell.style = "Hyperlink"  

        workbook.save(path)

    @staticmethod
    def adjust_column_widths(path:str):
        workbook = openpyxl.load_workbook(path)
        for sheet in workbook.worksheets:
            for col in sheet.columns:
                max_length = 0
                column = None
                
                for cell in col:
                    if not isinstance(cell, openpyxl.cell.MergedCell):
                        column = cell.column_letter  # Get the column letter
                        break

                if column is None:
                    continue  # Skip if no column letter was found

                for cell in col:
                    if not isinstance(cell, openpyxl.cell.MergedCell):
                        try:
                            if cell.value is not None:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                
                adjusted_width = max_length + 2  # Adding some extra space for aesthetics
                sheet.column_dimensions[column].width = adjusted_width
        
        workbook.save(path)
            
        

    
    def create_table_t1_db(self, path_db:str):
        # CREATE T1
        self._create_table(primary_key_name=self.DATA_STRUCTURE_VAR.PRIMARY_KEY_T1,
                           primary_key_type=self.DATA_STRUCTURE_VAR.TYPE_PRIMARY_KEY_T1,
                           columns_data=self.DICT_DATA_EXTRUCTURE_T1,
                           name_table=self.DATA_STRUCTURE_VAR.T1_NAME,
                           path=path_db)
        
    
    
    def create_table_t2_db(self, path_db:str):
        
        self._create_table(primary_key_name=self.DATA_STRUCTURE_VAR.PRIMARY_KEY_T2, 
                           primary_key_type=self.DATA_STRUCTURE_VAR.TYPE_PRIMARY_KEY_T2,
                           columns_data=self.dict_model_t2, 
                           name_table=self.DATA_STRUCTURE_VAR.T2_NAME,
                           reference_table=self.DATA_STRUCTURE_VAR.T1_NAME,
                           foreign_key=self.DATA_STRUCTURE_VAR.SHARED_FOREIGN_KEY_T1_T2,
                           shared_foreign_key_is_primary_key=self.DATA_STRUCTURE_VAR.IS_SHARED_FOREIGN_KEY_IN_T2_PRIMARY_KEY,
                           path=path_db)
        
    
    def create_table_tx_db(self, path):
        self.create_database_by_data_frame_model(df_model_with_one_row=self.data_frame_model_tx_one_row,
                                                    shared_foreign_key_is_primary_key=self.DATA_STRUCTURE_VAR.IS_SHARED_FOREIGN_KEY_IN_TX_PRIMARY_KEY,
                                                    tx_default_table_name=self.DATA_STRUCTURE_VAR.TX_DEFAULT_NAME,
                                                    reference_table=self.DATA_STRUCTURE_VAR.T2_NAME,
                                                    path=path)
    
    @ct_log.count_ms
    def load_data_t1_db(self, path, overwrite:bool=False, check_integrity_data:bool= True, unique_cols:list[str] = None):
        
        self._load_data_frame_in_table(data_frame=self.data_frame_t1,
                                       name_table=self.DATA_STRUCTURE_VAR.T1_NAME,
                                       overwrite=overwrite,
                                       check_foreign_keys=check_integrity_data,
                                       unique_cols=unique_cols,
                                       path=path)

    @ct_log.count_ms
    def load_data_t2_db(self, path, overwrite:bool=False, check_integrity_data:bool= True,  unique_cols:list[str] = None):
        self._load_data_frame_in_table(data_frame=self.data_frame_t2,
                                       name_table=self.DATA_STRUCTURE_VAR.T2_NAME,
                                       overwrite=overwrite,
                                       check_foreign_keys=check_integrity_data,
                                       unique_cols=unique_cols,
                                       path=path)

    @ct_log.count_ms
    def load_data_tx_db(self, path:str, overwrite = False):
         self._load_data_frame_multi_index(data_frame=self.data_frame_tx,
                                             tx_default_table_name=self.DATA_STRUCTURE_VAR.TX_DEFAULT_NAME,
                                             path=path,
                                             overwrite=overwrite)

        
    @staticmethod
    def get_sqlite_type(value):
        if isinstance(value, int):
            return SqlType.INTEGER.value
        elif isinstance(value, float):
            return SqlType.REAL.value
        elif isinstance(value, str):
            return SqlType.TEXT.value
        else:
            return SqlType.TEXT.value
    
    def _is_reserved_word(self, word: str) -> bool:
        return word.upper() in SQLITE_RESERVED_WORDS

    def _sanitize_column_name(self, column_name: str) -> str:
        if self._is_reserved_word(column_name):
            return f'"{column_name}"'
        return column_name
        
    def _create_table(self,
                    primary_key_name: str | None = None,
                    primary_key_type: SqlType | None =None,
                    columns_data: dict[str, any] = {},
                    name_table: str = "PROJECT",
                    foreign_key: str | None = None,
                    reference_table: str | None = None,
                    auto_increment_primary_key:bool= False,
                    shared_foreign_key_is_primary_key:bool= False, 
                    path: str = 'test_table_db.db',
                    overwrite: bool = True):  
        
        conn = sqlite3.connect(path)
        cursor = conn.cursor()

        if overwrite:
            cursor.execute(f"DROP TABLE IF EXISTS {name_table}")

        columns_definitions = []
        columns_data = columns_data.copy()
        columns_data.pop(primary_key_name, None)
        
        if auto_increment_primary_key:
            columns_definitions.append("id INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL")
        
        elif shared_foreign_key_is_primary_key and foreign_key and reference_table:
            sanitized_primary_key = self._sanitize_column_name(foreign_key)
            columns_definitions.append(f"{sanitized_primary_key} PRIMARY KEY")
        
        elif primary_key_name:
            sanitized_primary_key = self._sanitize_column_name(primary_key_name)
            columns_definitions.append(f"{sanitized_primary_key} {primary_key_type.value} PRIMARY KEY NOT NULL")
        
        for column_name, value in columns_data.items():
            sanitized_column_name = self._sanitize_column_name(column_name)
            sql_type = self.get_sqlite_type(value)
            columns_definitions.append(f"{sanitized_column_name} {sql_type}")

        if foreign_key and reference_table:
            sanitized_foreign_key = self._sanitize_column_name(foreign_key)
            columns_definitions.append(f"FOREIGN KEY ({sanitized_foreign_key}) REFERENCES {reference_table} ({sanitized_foreign_key} ) ON DELETE CASCADE")
        
        columns_definitions_str = ",\n                ".join(columns_definitions)

        cursor.execute(f'''
            CREATE TABLE {name_table} (
                {columns_definitions_str}
            )
        ''')

        conn.commit()
        conn.close()
        
        
        
    def _build_data_frame_from_dict(self, dictionary_data: dict[str, str]):
        # Create DataFrame
        data_frame = pd.DataFrame([dictionary_data])
        
        return data_frame
        

    def _load_data_frame_in_table(self, 
                                data_frame: pd.DataFrame,
                                unique_cols: list = None,  
                                overwrite: bool = False, 
                                check_foreign_keys: bool = True,
                                name_table: str = "PROJECT",
                                path: str = 'test_table_db.db'):
        conn = sqlite3.connect(path)
        
        if check_foreign_keys:
            cursor = conn.cursor()
            cursor.execute('PRAGMA foreign_keys = ON;')
        
        if overwrite:
            data_frame.to_sql(f'{name_table}', conn, if_exists='replace', index=False)
        else:
            if unique_cols:
                columns_str = ', '.join(unique_cols)
                existing_ids_query = f"SELECT {columns_str} FROM {name_table}"
                existing_ids_df = pd.read_sql_query(existing_ids_query, conn)
                
                if not existing_ids_df.empty:
                    existing_ids_set = set(tuple(row) for row in existing_ids_df.values)
                    filtered_data_frame = data_frame[~data_frame[unique_cols].apply(tuple, axis=1).isin(existing_ids_set)]
                else:
                    filtered_data_frame = data_frame
            else:
                filtered_data_frame = data_frame
            
            filtered_data_frame.to_sql(f'{name_table}', conn, if_exists='append', index=False)
        
        conn.commit()
        conn.close()
        
        
    def _load_data_frame_multi_index(   self,
                                        data_frame: pd.DataFrame,  
                                        tx_default_table_name: str = "TX",
                                        undefined="",
                                        path='test_table_db.db',
                                        overwrite= False,
                                        
                                        ):
        common_columns = None
        common_data:pd.DataFrame = None
        table_names = set()
        
        def _get_type_with_first_value(table_data, column):
            first_value = table_data[column].dropna().iloc[0] if not table_data[column].dropna().empty else ''
            col_type = self.get_sqlite_type(first_value)
            return col_type
        
        with sqlite3.connect(path) as conn:
            cursor = conn.cursor()
            
            for table_name in data_frame.columns.levels[0]:
                table_data = data_frame[table_name].copy()
                
                if data_frame.columns.nlevels > 2:
                    table_data.columns = ['_'.join(col) for col in table_data.columns]
                    table_data.columns = [col.replace('_value', '') for col in  table_data.columns]       
                                     
                if table_name == undefined:
                    common_columns = table_data.columns
                    common_data = table_data.copy() 
                else:
                    table_names.add(table_name)

            if table_names:
                for table_name in table_names:
                    table_data = data_frame[table_name].copy()  # Hacer una copia explícita del DataFrame

                    if data_frame.columns.nlevels > 2:
                        table_data.columns = ['_'.join(col) for col in table_data.columns]
                        table_data.columns = [col.replace('_value', '') for col in  table_data.columns]                        

                    # Add common columns if they exist
                    if common_columns is not None:
                        for col in common_columns:
                            if col not in table_data.columns:
                                table_data[col] = common_data[col].reindex(table_data.index, fill_value=None)
                                
                    existing_columns = [row[1] for row in cursor.execute(f'PRAGMA table_info("{table_name}")')]
                        
                    for col in table_data.columns:
                        if col not in existing_columns:
                            col_type = _get_type_with_first_value(table_data=table_data, column=col)
                            sanitized_col = self._sanitize_column_name(col)
                            cursor.execute(f'ALTER TABLE "{table_name}" ADD COLUMN {sanitized_col} {col_type}')

    
                    placeholders = ', '.join(['?'] * (len(table_data.columns) + 1)) 
                    data_to_insert = [(index, *row) for index, row in table_data.iterrows()]
                    if overwrite:
                        sql_insert_sentence = f'INSERT OR REPLACE INTO "{table_name}" VALUES ({placeholders})'
                    else:
                        sql_insert_sentence = f'INSERT OR IGNORE INTO "{table_name}" VALUES ({placeholders})'
                    cursor.executemany(sql_insert_sentence, data_to_insert)
                        
    
            else:
                if common_columns is not None:
                    placeholders = ', '.join(['?'] * (len(common_data.columns) + 1))  
                    data_to_insert = [(index, *row) for index, row in common_data.iterrows()]
                    if overwrite:
                        sql_insert_sentence = f'INSERT OR REPLACE INTO "{tx_default_table_name}" VALUES ({placeholders})'
                    else:
                        sql_insert_sentence = f'INSERT OR IGNORE INTO "{tx_default_table_name}" VALUES ({placeholders})'
                    cursor.executemany(sql_insert_sentence, data_to_insert)
                    

    def create_database_by_data_frame_model(self,
                                        df_model_with_one_row: pd.DataFrame,  
                                        auto_increment_primary_key:bool= False,
                                        shared_foreign_key_is_primary_key:bool= True, 
                                        reference_table:str | None = None,
                                        overwrite = True,
                                        tx_default_table_name: str = "TX",
                                        undefined="",
                                        path='test_table_db.db'):
        
        # Description:
        # if an element does not belong to any specific Tx, 
        # it is considered to belong to the union of all Tx×. 
        # In this representation:
        # - U is the universal set containing all possible elements.
        # - {TX1, TX2, ..., TXn} represents specific Table.
        # - FIRST_DIMENSIONAL_VALUE is the element that initially does not belong to any specific set.
        # - When FIRST_DIMENSIONAL_VALUE is not present, it is considered part of all Tx×.
        #
        # Visualization:
        #
        # U
        # +-------------------------+
        # | TX1   TX2   TX3  ...   TXn|
        # |   +--+  +--+  +--+  +--+  |
        # |   |  |  |  |  |  |  |  |  |
        # |   +--+  +--+  +--+  +--+  |
        # |                           |
        # |   FIRST_DIMENSIONAL_VALUE |
        # |   (added to all sets)     |
        # +-------------------------+
        foreign_key = None

        for table_name in df_model_with_one_row.columns.levels[0]: 
            table_data = df_model_with_one_row[table_name]
            
            foreign_key = table_data.index[0]
            break 
        
        def _add_primary_key() -> list[str]:
            columns_definitions = []

            if shared_foreign_key_is_primary_key:
                sanitized_primary_key = self._sanitize_column_name(foreign_key)
                columns_definitions.append(f"{sanitized_primary_key} PRIMARY KEY")
            elif auto_increment_primary_key:
                columns_definitions.append("id INTEGER PRIMARY KEY AUTOINCREMENT")
            return columns_definitions

        def _get_type_with_first_value(table_data, column):
            first_value = table_data[column].dropna().iloc[0] if not table_data[column].dropna().empty else ''
            col_type = self.get_sqlite_type(first_value)
            return col_type
    
        def _create_columns(table_data:pd.DataFrame = None,
                            columns_definitions:list[str] = None,
                            columns:tuple[str]|None=None) -> list[str]:
            
            if columns is not None:  # Handling the case when only column names are passed
                for col in columns:
                    col_type = _get_type_with_first_value(table_data=table_data, column=col) 
                    sanitized_col = self._sanitize_column_name(col)
                    columns_definitions.append(f'"{sanitized_col}" {col_type}')
                return columns_definitions

            elif table_data is not None:  # Original behavior
                for col in table_data.columns:
                    col_type = _get_type_with_first_value(table_data=table_data, column=col)
                    sanitized_col = self._sanitize_column_name(col)
                    columns_definitions.append(f'"{sanitized_col}" {col_type}')
            
                return columns_definitions
        
        def _add_foreign_key(columns_definitions:list[str]) -> list[str]:
            if foreign_key and reference_table:
                sanitized_foreign_key = self._sanitize_column_name(foreign_key)
                columns_definitions.append(f"FOREIGN KEY ({sanitized_foreign_key}) REFERENCES {reference_table} ({sanitized_foreign_key}) ON DELETE CASCADE")
            return columns_definitions
        
        def _create_table_sql_sentence(table_name, columns_definitions:list[str]) -> str :
            columns_definitions_str = ",\n                ".join(columns_definitions)
            string_sql = f'CREATE TABLE IF NOT EXISTS "{table_name}" ({columns_definitions_str});'
            return string_sql
        
        
        with sqlite3.connect(path) as conn:
            cursor = conn.cursor()

            # Initialize a dictionary to store common columns (from 'undefined')
            common_columns = None
            common_data:pd.DataFrame = None
            table_names = set()

            # Iterate over different levels of the MultiIndex to create tables
            for table_name in df_model_with_one_row.columns.levels[0]:  # The first level of the MultiIndex represents the tables
                # Extract data from the corresponding "table"
                table_data = df_model_with_one_row[table_name].copy()  # Explicitly make a copy of the DataFrame
                
                # Flatten the MultiIndex to get column names in a usable format
                if df_model_with_one_row.columns.nlevels > 2:
                    table_data.columns = ['_'.join(col) for col in table_data.columns]
                    table_data.columns = [col.replace('_value', '') for col in  table_data.columns]       

                # If the table name is 'undefined', we store it as a common column
                if table_name == undefined:
                    common_columns = table_data.columns
                    common_data = table_data.copy()  
                else:
                    table_names.add(table_name)

            if table_names:
                
                for table_name in table_names:
                    table_data = df_model_with_one_row[table_name].copy()  # Explicitly make a copy of the DataFrame

                    # Flatten the MultiIndex to get column names in a usable format only if greater than 2
                    if df_model_with_one_row.columns.nlevels > 2:
                        table_data.columns = ['_'.join(col) for col in table_data.columns]
                        table_data.columns = [col.replace('_value', '') for col in  table_data.columns]       

                    if common_columns is not None and len(common_columns) > 0:
                        for col in common_columns:
                            if col not in table_data.columns:
                                table_data[col] = common_data[col].iloc[0]  # or handle differently if necessary

                    columns_definitions = _add_primary_key()
                    columns_definitions = _create_columns(table_data=table_data, columns_definitions=columns_definitions)
                    columns_definitions = _add_foreign_key(columns_definitions=columns_definitions)
                    string_sql = _create_table_sql_sentence(table_name=table_name, columns_definitions=columns_definitions)
                    if overwrite:
                        cursor.execute(f"DROP TABLE IF EXISTS {table_name}")
                    cursor.execute(string_sql)

                    existing_columns = [row[1] for row in cursor.execute(f'PRAGMA table_info("{table_name}")')]
                    
                    for col in table_data.columns:
                        if col not in existing_columns:
                            col_type = _get_type_with_first_value(table_data=table_data, column=col)
                            sanitized_col = self._sanitize_column_name(col)
                            cursor.execute(f'ALTER TABLE "{table_name}" ADD COLUMN {sanitized_col} {col_type}')

            else:
                
                if common_columns is not None and len(common_columns) > 0:
                    if  len(common_columns) != len(set(common_columns)):
                        raise DuplicatedColumnNames(message="The columns have duplicated names!")
                    columns_definitions = _add_primary_key()
                    common_data = common_data.sort_index()
                    columns_definitions = _create_columns(table_data=common_data ,columns=common_columns, columns_definitions=columns_definitions)
                    columns_definitions = _add_foreign_key(columns_definitions=columns_definitions)
                    if overwrite:
                        cursor.execute(f"DROP TABLE IF EXISTS {tx_default_table_name}")
                    string_sql = _create_table_sql_sentence(table_name=tx_default_table_name, columns_definitions=columns_definitions)
                    cursor.execute(string_sql)

if __name__ == "__main__":
    
# o.OOOo.                       .oOOOo.                                                            oOoOOoOOo                  
# O    `o                      o     o                                                                o                      
# o      O          O          O.         O                         O                                 o                   O  
# O      o         oOo          `OOoo.   oOo                       oOo                                O                  oOo 
# o      O .oOoO'   o   .oOoO'       `O   o   `OoOo. O   o  .oOo    o   O   o  `OoOo. .oOo.           o     .oOo. .oOo    o  
# O      o O   o    O   O   o         o   O    o     o   O  O       O   o   O   o     OooO'           O     OooO' `Ooo.   O  
# o    .O' o   O    o   o   O  O.    .O   o    O     O   o  o       o   O   o   O     O               O     O         O   o  
# OooOO'   `OoO'o   `oO `OoO'o  `oooO'    `oO  o     `OoO'o `OoO'   `oO `OoO'o  o     `OoO'           o'    `OoO' `OoO'   `oO
# pebbles                                                                                                                                               
# Description:
# This test is incomplete Sep 3, but working
# 
# Usage:
# python data_structure.py
# 
# Author: [Denis Bueltan]
# Date: [Sep 2]
# last change [Sep 6]
# if the something change this hash will do. 
# original_hash_test_table_db = "dd3b636b1ca7422384ff8ae5caebba53ea715c29415867ee2e7b58e6a7b97592"                                                                                                               
# =============================================================================================================================
    

    _data_t1 = {
                'screen_id': "0_fac_09-09-2024_ift4_video_source_video",
                'name': "Test_1",
                'source_type': "VIDEO",
                'source': "FULL FLIGHT! SpaceX Starship Flight 4.mp4",
                'resolution': "FULL_HD",
                
            }
    
#     _list_data_example = [
#     {   'screen_id':"0_gcy_05-09-2024_ift4_vide_file_web",
#         'image_name': 'IFP_21_1725031126.3294787_change_detected.png',
#         'image_capture_data': {
#             'session_id': '1725031110708_1725031110_ift4_starship_IFP_session',
#             'session_code': 'IFP',
#             'is_start_session': False,
#             'is_end_session': False,
#             'is_change_detected': True,
#             'timestamp_seconds': 15.973,
#             'session_captures_number': 21,
#             'date': 'Fri Aug 30 2024 15:18:46 GMT',
#         },
#         'areas_ocr': [
#             {'area_number': 2, 'name': 'ALTITUDE', 'group': "STAGE_1", 'value': 11, 'raw_value': '11', 'path_image': 'ProjectFiles/ift4_starship/D_CROPPED_IMAGES_OCR/3_agf_30-08-2024_ift4_starship_web/source_IFP_21_1725031126.3294787_change_detected_area_no_2.png'},
#             {'area_number': 3, 'name': 'SPEED', 'group': "STAGE_1", 'value': 201, 'raw_value': '201', 'path_image': 'ProjectFiles/ift4_starship/D_CROPPED_IMAGES_OCR/3_agf_30-08-2024_ift4_starship_web/source_IFP_21_1725031126.3294787_change_detected_area_no_3.png'},
#             {'area_number': 0, 'name': 'ALTITUDE', 'group': "STAGE_2", 'value': 12, 'raw_value': '12', 'path_image': 'ProjectFiles/ift4_starship/D_CROPPED_IMAGES_OCR/3_agf_30-08-2024_ift4_starship_web/source_IFP_21_1725031126.3294787_change_detected_area_no_0.png'},
#             {'area_number': 4, 'name': 'SPEED', 'group': "STAGE_2", 'value': 202, 'raw_value': '202', 'path_image': 'ProjectFiles/ift4_starship/D_CROPPED_IMAGES_OCR/3_agf_30-08-2024_ift4_starship_web/source_IFP_21_1725031126.3294787_change_detected_area_no_4.png'},
#             {'area_number': 1, 'name': 'TIME_LAPSE', 'group': None, 'value': 'T+00:00:12', 'raw_value': 'T+00:00:12', 'path_image': 'ProjectFiles/ift4_starship/D_CROPPED_IMAGES_OCR/3_agf_30-08-2024_ift4_starship_web/source_IFP_21_1725031126.3294787_change_detected_area_no_1.png'}
#         ]
#     },
    
#     {   'screen_id':"0_gcy_05-09-2024_ift4_vide_file_web",
#         'image_name': 'IFP_22_1725031126.3294787_change_detected.png',
#         'image_capture_data': {
#             'session_id': '1725031110708_1725031110_ift4_starship_IFP_session',
#             'session_code': 'ZZZ',
#             'is_start_session': False,
#             'is_end_session': False,
#             'is_change_detected': True,
#             'timestamp_seconds': 20.973,
#             'session_captures_number': 22,
#             'date': 'Fri Aug 31 2024 12:12:46 GMT',
#         },
#         'areas_ocr': [
#             {'area_number': 2, 'name': 'ALTITUDE', 'group': "STAGE_1", 'value': 21, 'raw_value': '21', 'path_image': 'ProjectFiles/ift4_starship/D_CROPPED_IMAGES_OCR/3_agf_30-08-2024_ift4_starship_web/source_IFP_21_1725031126.3294787_change_detected_area_no_2.png'},
#             {'area_number': 3, 'name': 'SPEED', 'group': "STAGE_1", 'value': 401, 'raw_value': '401', 'path_image': 'ProjectFiles/ift4_starship/D_CROPPED_IMAGES_OCR/3_agf_30-08-2024_ift4_starship_web/source_IFP_21_1725031126.3294787_change_detected_area_no_3.png'},
#             {'area_number': 0, 'name': 'ALTITUDE', 'group': "STAGE_2", 'value': 22, 'raw_value': '22', 'path_image': 'ProjectFiles/ift4_starship/D_CROPPED_IMAGES_OCR/3_agf_30-08-2024_ift4_starship_web/source_IFP_21_1725031126.3294787_change_detected_area_no_0.png'},
#             {'area_number': 4, 'name': 'SPEED', 'group': "STAGE_2", 'value': 402, 'raw_value': '402', 'path_image': 'ProjectFiles/ift4_starship/D_CROPPED_IMAGES_OCR/3_agf_30-08-2024_ift4_starship_web/source_IFP_21_1725031126.3294787_change_detected_area_no_4.png'},
#             {'area_number': 1, 'name': 'TIME_LAPSE', 'group': None, 'value': 'T+00:00:13', 'raw_value': 'T+00:00:13', 'path_image': 'ProjectFiles/ift4_starship/D_CROPPED_IMAGES_OCR/3_agf_30-08-2024_ift4_starship_web/source_IFP_21_1725031126.3294787_change_detected_area_no_1.png'}
#         ]
#     },
#    # Add additional entries here...
# ]   
    _list_data_example = [
        
        {'screen_id': '0_fac_09-09-2024_ift4_video_source_video',
         'image_name': 'SJP_694_1725895501_change_detected',
         'image_capture_data': {'session_id': '1725895398513_1725895398_ift4_video_source_SJP_session', 
                                'session_code': 'SJP',
                                'is_start_session': False,
                                'is_end_session': False,
                                'is_change_detected': True,
                                'timestamp_seconds': 103.183,
                                'session_captures_number': 694, 'date': 'Mon Sep 09 2024 15:25:01 GMT'},
         'areas_ocr': [{'area_number': 100, 'path_image': '--', 'name': 'ALTITUDE', 'group': 'STAGE_2', 'raw_value': '--', 'value': 100},
                       {'area_number': 100, 'path_image': '--', 'name': 'ALTITUDE', 'group': 'STAGE_1', 'raw_value': '--', 'value': 100},
                       {'area_number': 100, 'path_image': '--', 'name': 'TIME_LAPSE', 'group': None, 'raw_value': '--', 'value': '--'}, 
                       {'area_number': 100, 'path_image': '--', 'name': 'SPEED', 'group': 'STAGE_2', 'raw_value': '--', 'value': 100}, 
                       {'area_number': 100, 'path_image': '--', 'name': 'SPEED', 'group': 'STAGE_1', 'raw_value': '--', 'value': 100}]},
        {'screen_id': '0_fac_09-09-2024_ift4_video_source_video',
         'image_name': 'SJP_695_1725895501_change_detected',
         'image_capture_data': {'session_id': '1725895398513_1725895398_ift4_video_source_SJP_session', 
                                'session_code': 'SJP',
                                'is_start_session': False,
                                'is_end_session': False, 
                                'is_change_detected': True,
                                'timestamp_seconds': 103.306,
                                'session_captures_number': 695, 
                                'date': 'Mon Sep 09 2024 15:25:01 GMT'}, 
         'areas_ocr': [{'area_number': 1, 'name': 'ALTITUDE', 'group': 'STAGE_1', 'value': None, 'raw_value': None, 'path_image': 'Projects/ift4_video_source/D_CROPPED_IMAGES_OCR/0_fac_09-09-2024_ift4_video_source_video/source_SJP_695_1725895501_change_detected_area_no_1.png'},
                       {'area_number': 4, 'name': 'SPEED', 'group': 'STAGE_1', 'value': None, 'raw_value': None, 'path_image': 'Projects/ift4_video_source/D_CROPPED_IMAGES_OCR/0_fac_09-09-2024_ift4_video_source_video/source_SJP_695_1725895501_change_detected_area_no_4.png'}, 
                       {'area_number': 0, 'name': 'ALTITUDE', 'group': 'STAGE_2', 'value': 162, 'raw_value': '162', 'path_image': 'Projects/ift4_video_source/D_CROPPED_IMAGES_OCR/0_fac_09-09-2024_ift4_video_source_video/source_SJP_695_1725895501_change_detected_area_no_0.png'},
                       {'area_number': 3, 'name': 'SPEED', 'group': 'STAGE_2', 'value': 26450, 'raw_value': '26450', 'path_image': 'Projects/ift4_video_source/D_CROPPED_IMAGES_OCR/0_fac_09-09-2024_ift4_video_source_video/source_SJP_695_1725895501_change_detected_area_no_3.png'},
                       {'area_number': 2, 'name': 'TIME_LAPSE', 'group': None, 'value': 'T+00:10:19', 'raw_value': 'T+00:10:19', 'path_image': 'Projects/ift4_video_source/D_CROPPED_IMAGES_OCR/0_fac_09-09-2024_ift4_video_source_video/source_SJP_695_1725895501_change_detected_area_no_2.png'}]}
        
    ]
   

    LIST_DATA_SOURCE = deepcopy(_list_data_example)
    KEY_LIST_TABLE_TX:str = "areas_ocr"
    PRIMARY_KEY_T1:str = "screen_id"
    PRIMARY_KEY_T2:str = "image_name"
 
    
    data_structure_var = DataStructureVar(
    DICT_DATA_EXTRUCTURE_T1=_data_t1,
    DICT_DATA_EXTRUCTURE_T2_TX=_list_data_example[0],
    
    T1_NAME = "PROJECT",
    PRIMARY_KEY_T1 = PRIMARY_KEY_T1,
    TYPE_PRIMARY_KEY_T1= SqlType.TEXT,
    
    T2_NAME = "COLLECTOR",
    PRIMARY_KEY_T2 = PRIMARY_KEY_T2,
    TYPE_PRIMARY_KEY_T2= SqlType.TEXT,
    
    TX_DEFAULT_NAME = _data_t1['name'],
    KEY_LIST_TABLE_TX = KEY_LIST_TABLE_TX,
    #KEYS_TO_UNPACK_IN_TX is some key not present will ignore 
    KEYS_TO_UNPACK_IN_TX = UnpackKeysToTX(KEYS=['timestamp_seconds', 'session_code',]) ,
    KEYS_TO_AVOID_UNPACK_IN_TX = AvoidUnpackKeysToTX(KEYS=[KEY_LIST_TABLE_TX]) ,
    TX_DUAL_INDEX = TxDualIndex(INDEX_0="group",INDEX_1="name"),
    #TX_THIRD_INDEX_DIMENSION is some key not present will ignore 
    TX_THIRD_INDEX_DIMENSION=TxThirdIndexDimension(LIST_INDEX=["value", "raw_value"]),
    KEYS_TO_UNPACK_IN_DATA_DICT_T2 =  UnpackKeysToT2(KEYS=["image_capture_data"]) ,
    KEYS_TO_AVOID_UNPACK_DATA_DICT_T2 = AvoidUnpackKeysToT2(KEYS=[KEY_LIST_TABLE_TX]),
    SHARED_FOREIGN_KEY_T1_T2 = PRIMARY_KEY_T1,
    SHARED_FOREIGN_KEY_T2_TX =  PRIMARY_KEY_T2,
    IS_SHARED_FOREIGN_KEY_IN_T2_PRIMARY_KEY= False,
    IS_SHARED_FOREIGN_KEY_IN_TX_PRIMARY_KEY= True,
  )
    data_structure_var.validate_tx_dual_index()
    #print (data_structure_var)
    d_structure = DataStructure()

    # CREATE T1
    d_structure._create_table(primary_key_name=data_structure_var.PRIMARY_KEY_T1,
                             primary_key_type=data_structure_var.TYPE_PRIMARY_KEY_T1,
                             columns_data=_data_t1,
                             name_table=data_structure_var.T1_NAME)
    
    #DICT T2 USED TO CREATE THE TABLE AND DEFINE TYPES
    
    dict_model_t2 = d_structure._build_dictionary(dictionary_data=data_structure_var.DICT_DATA_EXTRUCTURE_T2_TX,
                                                keys_to_unpack=data_structure_var.KEYS_TO_UNPACK_IN_DATA_DICT_T2.KEYS,
                                                keys_to_remove=data_structure_var.KEYS_TO_AVOID_UNPACK_DATA_DICT_T2.KEYS)
    
    
    #CREATE TABLE T2 
    d_structure._create_table(primary_key_name=data_structure_var.PRIMARY_KEY_T2, 
                              primary_key_type=data_structure_var.TYPE_PRIMARY_KEY_T2,
                              columns_data=dict_model_t2, 
                              name_table=data_structure_var.T2_NAME,
                              reference_table=data_structure_var.T1_NAME,
                              foreign_key=data_structure_var.SHARED_FOREIGN_KEY_T1_T2)
    
    
    #- > LIST DICT USED TO CREATE DYNAMIC TX× . 
    tx_tables_dict = d_structure._filter_and_flattens_dict_from_keywords(input_dict=data_structure_var.DICT_DATA_EXTRUCTURE_T2_TX,
                                                                        keywords_to_get=[data_structure_var.KEY_LIST_TABLE_TX],
                                                                        keywords_to_ignore=data_structure_var.KEYS_TO_UNPACK_IN_DATA_DICT_T2.KEYS)
    
    
    dict_to_include_in_txs = d_structure._filter_and_flattens_dict_from_keywords(input_dict=data_structure_var.DICT_DATA_EXTRUCTURE_T2_TX,
                                                                                 keywords_to_get= data_structure_var.KEYS_TO_UNPACK_IN_TX.KEYS,
                                                                                 keywords_to_ignore= data_structure_var.KEYS_TO_AVOID_UNPACK_IN_TX.KEYS)
    
    
    
   
    list_model_to_add_tx = d_structure._create_list_to_extend_in_tx(dict_to_include_in_txs=dict_to_include_in_txs, 
                                                              tx_dual_index=data_structure_var.TX_DUAL_INDEX,
                                                              tx_third_index_dimension=data_structure_var.TX_THIRD_INDEX_DIMENSION)
    
    tx_tables_model:list[dict] = tx_tables_dict[KEY_LIST_TABLE_TX]
    # INSERT OPTIONAL COLUMNS IN TX .  
    tx_tables_model.extend(list_model_to_add_tx)
    
    model_data_frame_tx_one_row = d_structure._create_multi_index_data_frame_model(list_dict_model=tx_tables_model,
                                                                     tx_dual_index=data_structure_var.TX_DUAL_INDEX,
                                                                     third_dimension_index_keys=data_structure_var.TX_THIRD_INDEX_DIMENSION,
                                                                     shared_foreign_key=data_structure_var.SHARED_FOREIGN_KEY_T2_TX,
                                                                     )
    
    d_structure.create_database_by_data_frame_model(df_model_with_one_row=model_data_frame_tx_one_row,
                                                    shared_foreign_key_is_primary_key=data_structure_var.IS_SHARED_FOREIGN_KEY_IN_TX_PRIMARY_KEY,
                                                    tx_default_table_name=data_structure_var.TX_DEFAULT_NAME,
                                                    reference_table=data_structure_var.T2_NAME)
    
       
    
    
    
    data_frame_t1 = d_structure._build_data_frame_from_dict(_data_t1)
    d_structure._load_data_frame_in_table(data_frame=data_frame_t1)
    
    
    data_frame_t2 = d_structure.create_data_frame_from_model(list_data=LIST_DATA_SOURCE,
                                                             keywords_to_get=list(dict_model_t2.keys()))
    
    d_structure._load_data_frame_in_table(data_frame=data_frame_t2, name_table=data_structure_var.T2_NAME)
    
    list_shared_keys_ts_tx = data_frame_t2[PRIMARY_KEY_T2].tolist()
    
    super_list_tx = [d_structure._extend_list_columns_txs(entry_dict=entry,
                                                          key_list_table_tx=KEY_LIST_TABLE_TX,
                                                          keywords_to_get=data_structure_var.KEYS_TO_UNPACK_IN_TX.KEYS,
                                                          keywords_to_ignore=data_structure_var.KEYS_TO_AVOID_UNPACK_IN_TX.KEYS,
                                                          tx_dual_index=data_structure_var.TX_DUAL_INDEX,
                                                          tx_third_index_dimension=data_structure_var.TX_THIRD_INDEX_DIMENSION,
                                                          ) for entry in LIST_DATA_SOURCE]

    model_data_frame_tx_empty = d_structure._create_multi_index_data_frame_model(list_dict_model=tx_tables_model,
                                                                    tx_dual_index=data_structure_var.TX_DUAL_INDEX,
                                                                    third_dimension_index_keys=data_structure_var.TX_THIRD_INDEX_DIMENSION,
                                                                    shared_foreign_key=data_structure_var.SHARED_FOREIGN_KEY_T2_TX,
                                                                    with_data=False,
                                                                    )

    data_frame_tx_with_data = d_structure._populate_data_frame_tx(  super_list_data=super_list_tx,
                                                                    df_model=model_data_frame_tx_empty,
                                                                    tx_dual_index=data_structure_var.TX_DUAL_INDEX,
                                                                    third_dimension_index_keys=data_structure_var.TX_THIRD_INDEX_DIMENSION,
                                                                    list_keys=list_shared_keys_ts_tx)
    
    
    
    
    
    print(data_frame_tx_with_data)
    d_structure._load_data_frame_multi_index(data_frame=data_frame_tx_with_data,
                                             tx_default_table_name=data_structure_var.TX_DEFAULT_NAME)

    
    # if something change this has will do .  
    last_hash_test_table_db = "dd3b636b1ca7422384ff8ae5caebba53ea715c29415867ee2e7b58e6a7b97592"
    

    
    
    current_has_db = hash_database(database_path="test_table_db.db")
    print("________CURRENT__HASH_________")
    print(f"{current_has_db}")
    print("________IT IS GOOD, AT LEAST KNOWN__HASH_________")
    print(f"{last_hash_test_table_db}")